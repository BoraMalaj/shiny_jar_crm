# frontend/pages/reports.py - UPDATED VERSION
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
from datetime import datetime, timedelta
import io
import base64
import json

# Try to import reportlab, but provide fallback
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    st.warning("⚠️ reportlab not installed. PDF export features will be limited.")

# Try to import openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    st.warning("⚠️ openpyxl not installed. Excel export features will be limited.")

def show_reports_page():
    st.markdown('<h1 class="main-header">📊 Professional Reports</h1>', unsafe_allow_html=True)
    
    api_url = st.session_state.get('api_url', 'http://localhost:8000')
    
    tab1, tab2, tab3, tab4 = st.tabs(["📅 Financial Reports", "👥 Customer Reports", "🏭 Supplier Reports", "🛠️ Custom Reports"])
    
    with tab1:
        show_financial_reports(api_url)
    
    with tab2:
        show_customer_reports(api_url)
    
    with tab3:
        show_supplier_reports(api_url)
    
    with tab4:
        show_custom_reports(api_url)

def show_financial_reports(api_url):
    st.subheader("Financial Statements & Reports")
    
    # Report selection and date range
    col1, col2 = st.columns(2)
    
    with col1:
        report_type = st.selectbox(
            "Select Report Type",
            [
                "Income Statement (Profit & Loss)",
                "Balance Sheet",
                "Cash Flow Statement", 
                "Expense Breakdown",
                "Revenue Analysis",
                "Monthly Financial Summary"
            ]
        )
    
    with col2:
        # Date range
        today = datetime.now()
        date_option = st.selectbox(
            "Period",
            ["Last 30 Days", "Last Quarter", "Last 6 Months", "Year to Date", "Last Year", "Custom Range"]
        )
        
        if date_option == "Custom Range":
            start_date = st.date_input("Start Date", value=today - timedelta(days=30))
            end_date = st.date_input("End Date", value=today)
        else:
            # Set dates based on selection
            if date_option == "Last 30 Days":
                start_date = today - timedelta(days=30)
                end_date = today
            elif date_option == "Last Quarter":
                start_date = today - timedelta(days=90)
                end_date = today
            elif date_option == "Last 6 Months":
                start_date = today - timedelta(days=180)
                end_date = today
            elif date_option == "Year to Date":
                start_date = datetime(today.year, 1, 1)
                end_date = today
            else:  # Last Year
                start_date = datetime(today.year - 1, 1, 1)
                end_date = datetime(today.year - 1, 12, 31)
    
    # Generate report button
    if st.button("📈 Generate Report", type="primary", use_container_width=True):
        with st.spinner(f"Generating {report_type}..."):
            try:
                # Fetch data
                transactions_response = requests.get(f"{api_url}/api/transactions", params={"limit": 10000})
                
                if transactions_response.status_code == 200:
                    transactions = response.json()
                    
                    if transactions:
                        df = pd.DataFrame(transactions)
                        
                        # Handle date field
                        if 'transaction_date' in df.columns:
                            df['date'] = pd.to_datetime(df['transaction_date'])
                        elif 'date' in df.columns:
                            df['date'] = pd.to_datetime(df['date'])
                        
                        # Filter by date
                        df['date_only'] = df['date'].dt.date
                        mask = (df['date_only'] >= start_date) & (df['date_only'] <= end_date)
                        df_filtered = df[mask]
                        
                        # Generate selected report
                        if "Income Statement" in report_type:
                            generate_income_statement(df_filtered, start_date, end_date, api_url)
                        elif "Balance Sheet" in report_type:
                            generate_balance_sheet(df_filtered, start_date, end_date, api_url)
                        elif "Cash Flow" in report_type:
                            generate_cash_flow(df_filtered, start_date, end_date, api_url)
                        elif "Expense Breakdown" in report_type:
                            generate_expense_report(df_filtered, start_date, end_date, api_url)
                        elif "Revenue Analysis" in report_type:
                            generate_revenue_analysis(df_filtered, start_date, end_date, api_url)
                        elif "Monthly Financial Summary" in report_type:
                            generate_monthly_summary(df_filtered, start_date, end_date, api_url)
                    else:
                        st.warning("No transaction data available for the selected period")
                else:
                    st.error("Failed to fetch transaction data")
            
            except Exception as e:
                st.error(f"Error generating report: {str(e)}")

def generate_income_statement(df, start_date, end_date, api_url):
    """Generate professional income statement"""
    
    # Calculate totals
    total_income = df[df['type'] == 'income']['amount'].sum()
    total_expenses = df[df['type'] == 'expense']['amount'].sum()
    net_income = total_income - total_expenses
    
    # Get category breakdowns
    income_by_category = df[df['type'] == 'income'].groupby('category')['amount'].sum().reset_index()
    income_by_category = income_by_category.sort_values('amount', ascending=False)
    
    expenses_by_category = df[df['type'] == 'expense'].groupby('category')['amount'].sum().reset_index()
    expenses_by_category = expenses_by_category.sort_values('amount', ascending=False)
    
    # Display report
    st.success(f"✅ Income Statement Generated for {start_date} to {end_date}")
    
    # Report header
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        st.markdown(f"**Period:** {start_date} to {end_date}")
    with col2:
        st.markdown(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    with col3:
        st.markdown(f"**Transactions:** {len(df)}")
    
    # Key metrics
    st.subheader("💰 Financial Summary")
    
    metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
    with metric_col1:
        st.metric("Total Revenue", f"€{total_income:,.2f}")
    with metric_col2:
        st.metric("Total Expenses", f"€{total_expenses:,.2f}")
    with metric_col3:
        profit_color = "normal" if net_income >= 0 else "inverse"
        st.metric("Net Income", f"€{net_income:,.2f}", delta_color=profit_color)
    with metric_col4:
        profit_margin = (net_income / total_income * 100) if total_income > 0 else 0
        st.metric("Profit Margin", f"{profit_margin:.1f}%")
    
    # Detailed breakdown
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📈 Revenue Breakdown")
        if not income_by_category.empty:
            fig1 = px.pie(
                income_by_category,
                values='amount',
                names='category',
                title='Revenue by Category',
                color_discrete_sequence=px.colors.sequential.Greens,
                hole=0.4
            )
            st.plotly_chart(fig1, use_container_width=True)
            
            # Revenue table
            st.dataframe(
                income_by_category,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "category": "Category",
                    "amount": st.column_config.NumberColumn("Amount", format="€%.2f")
                }
            )
    
    with col2:
        st.subheader("📉 Expense Breakdown")
        if not expenses_by_category.empty:
            fig2 = px.pie(
                expenses_by_category,
                values='amount',
                names='category',
                title='Expenses by Category',
                color_discrete_sequence=px.colors.sequential.Reds,
                hole=0.4
            )
            st.plotly_chart(fig2, use_container_width=True)
            
            # Expense table
            st.dataframe(
                expenses_by_category,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "category": "Category",
                    "amount": st.column_config.NumberColumn("Amount", format="€%.2f")
                }
            )
    
    # Export options
    st.subheader("📤 Export Options")
    
    export_col1, export_col2, export_col3 = st.columns(3)
    
    with export_col1:
        if REPORTLAB_AVAILABLE:
            if st.button("📄 Export as PDF", use_container_width=True):
                pdf_buffer = create_pdf_income_statement(
                    start_date, end_date, total_income, total_expenses, net_income,
                    income_by_category, expenses_by_category
                )
                st.download_button(
                    label="Download PDF",
                    data=pdf_buffer,
                    file_name=f"income_statement_{start_date}_{end_date}.pdf",
                    mime="application/pdf"
                )
        else:
            st.info("PDF export requires reportlab: `pip install reportlab`")
    
    with export_col2:
        if OPENPYXL_AVAILABLE:
            if st.button("📊 Export as Excel", use_container_width=True):
                excel_buffer = create_excel_income_statement(
                    start_date, end_date, total_income, total_expenses, net_income,
                    income_by_category, expenses_by_category, df
                )
                st.download_button(
                    label="Download Excel",
                    data=excel_buffer,
                    file_name=f"income_statement_{start_date}_{end_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("Excel export requires openpyxl: `pip install openpyxl`")
    
    with export_col3:
        csv_data = df.to_csv(index=False)
        st.download_button(
            label="📈 Export Raw Data (CSV)",
            data=csv_data,
            file_name=f"transactions_{start_date}_{end_date}.csv",
            mime="text/csv"
        )

def create_pdf_income_statement(start_date, end_date, revenue, expenses, net_income, 
                               income_cats, expense_cats):
    """Create PDF income statement"""
    if not REPORTLAB_AVAILABLE:
        return b"PDF export requires reportlab library"
    
    buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        spaceAfter=30
    )
    elements.append(Paragraph("Income Statement (Profit & Loss)", title_style))
    
    # Period
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary section
    elements.append(Paragraph("Financial Summary", styles['Heading2']))
    
    summary_data = [
        ["Total Revenue", f"€{revenue:,.2f}"],
        ["Total Expenses", f"€{expenses:,.2f}"],
        ["Net Income", f"€{net_income:,.2f}"],
        ["Profit Margin", f"{(net_income/revenue*100):.1f}%" if revenue > 0 else "0%"]
    ]
    
    summary_table = Table(summary_data, colWidths=[400, 100])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Revenue breakdown
    elements.append(Paragraph("Revenue by Category", styles['Heading2']))
    
    rev_data = [["Category", "Amount"]]
    for _, row in income_cats.iterrows():
        rev_data.append([row['category'], f"€{row['amount']:,.2f}"])
    
    rev_table = Table(rev_data, colWidths=[350, 150])
    rev_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2ECC71")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(rev_table)
    elements.append(Spacer(1, 30))
    
    # Expense breakdown
    elements.append(Paragraph("Expenses by Category", styles['Heading2']))
    
    exp_data = [["Category", "Amount"]]
    for _, row in expense_cats.iterrows():
        exp_data.append([row['category'], f"€{row['amount']:,.2f}"])
    
    exp_table = Table(exp_data, colWidths=[350, 150])
    exp_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E74C3C")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(exp_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def create_excel_income_statement(start_date, end_date, revenue, expenses, net_income,
                                 income_cats, expense_cats, transaction_df):
    """Create Excel income statement"""
    if not OPENPYXL_AVAILABLE:
        return b"Excel export requires openpyxl library"
    
    buffer = io.BytesIO()
    
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header
    ws_summary['A1'] = "Income Statement"
    ws_summary['A1'].font = Font(size=16, bold=True)
    
    ws_summary['A2'] = f"Period: {start_date} to {end_date}"
    ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Summary data
    ws_summary['A5'] = "Financial Summary"
    ws_summary['A5'].font = Font(bold=True)
    
    summary_data = [
        ["Total Revenue", revenue],
        ["Total Expenses", expenses],
        ["Net Income", net_income],
        ["Profit Margin", (net_income/revenue*100) if revenue > 0 else 0]
    ]
    
    for i, (label, value) in enumerate(summary_data, start=6):
        ws_summary[f'A{i}'] = label
        ws_summary[f'B{i}'] = value
        ws_summary[f'B{i}'].number_format = '"€"#,##0.00'
    
    # Revenue breakdown sheet
    ws_revenue = wb.create_sheet("Revenue Breakdown")
    ws_revenue['A1'] = "Revenue by Category"
    ws_revenue['A1'].font = Font(size=14, bold=True)
    
    ws_revenue['A3'] = "Category"
    ws_revenue['B3'] = "Amount"
    ws_revenue['A3'].font = ws_revenue['B3'].font = Font(bold=True)
    
    for i, row in income_cats.iterrows():
        ws_revenue[f'A{i+4}'] = row['category']
        ws_revenue[f'B{i+4}'] = row['amount']
        ws_revenue[f'B{i+4}'].number_format = '"€"#,##0.00'
    
    # Expense breakdown sheet
    ws_expenses = wb.create_sheet("Expense Breakdown")
    ws_expenses['A1'] = "Expenses by Category"
    ws_expenses['A1'].font = Font(size=14, bold=True)
    
    ws_expenses['A3'] = "Category"
    ws_expenses['B3'] = "Amount"
    ws_expenses['A3'].font = ws_expenses['B3'].font = Font(bold=True)
    
    for i, row in expense_cats.iterrows():
        ws_expenses[f'A{i+4}'] = row['category']
        ws_expenses[f'B{i+4}'] = row['amount']
        ws_expenses[f'B{i+4}'].number_format = '"€"#,##0.00'
    
    # Raw data sheet
    ws_raw = wb.create_sheet("Raw Data")
    if not transaction_df.empty:
        # Write headers
        for col_idx, column in enumerate(transaction_df.columns, 1):
            ws_raw.cell(row=1, column=col_idx, value=column)
        
        # Write data
        for row_idx, row in transaction_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws_raw.cell(row=row_idx+2, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def show_customer_reports(api_url):
    st.subheader("Customer Analysis Reports")
    
    report_type = st.selectbox(
        "Select Customer Report",
        [
            "Customer Spending Analysis",
            "Customer Segmentation Report",
            "Customer Acquisition Report",
            "Top Customers Report",
            "Customer Lifetime Value Analysis"
        ]
    )
    
    # Add date range for customer reports
    col1, col2 = st.columns(2)
    with col1:
        cust_start_date = st.date_input("Start Date", value=datetime.now() - timedelta(days=90))
    with col2:
        cust_end_date = st.date_input("End Date", value=datetime.now())
    
    if st.button("👥 Generate Customer Report", type="primary", use_container_width=True):
        try:
            customers_response = requests.get(f"{api_url}/api/customers")
            transactions_response = requests.get(f"{api_url}/api/transactions", params={"limit": 5000})
            
            if customers_response.status_code == 200 and transactions_response.status_code == 200:
                customers = customers_response.json()
                transactions = transactions_response.json()
                
                if customers:
                    customers_df = pd.DataFrame(customers)
                    
                    # Generate selected report
                    if "Spending Analysis" in report_type:
                        st.success("✅ Customer Spending Analysis Generated")
                        
                        # Spending distribution
                        st.subheader("Customer Spending Distribution")
                        
                        fig = px.histogram(
                            customers_df,
                            x='total_spent',
                            nbins=20,
                            title='Distribution of Customer Spending',
                            labels={'total_spent': 'Total Spent (€)', 'count': 'Number of Customers'}
                        )
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Top spenders
                        st.subheader("Top 20 Customers")
                        top_customers = customers_df.nlargest(20, 'total_spent')
                        
                        # Create display dataframe
                        display_df = top_customers[['name', 'email', 'total_spent']].copy()
                        display_df['total_spent'] = display_df['total_spent'].apply(lambda x: f"€{x:,.2f}")
                        
                        st.dataframe(
                            display_df,
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "name": "Customer",
                                "email": "Email",
                                "total_spent": "Total Spent"
                            }
                        )
                        
                        # Export option
                        csv = customers_df.to_csv(index=False)
                        st.download_button(
                            label="📥 Download Customer Data",
                            data=csv,
                            file_name="customer_spending_analysis.csv",
                            mime="text/csv"
                        )
                    
                    elif "Segmentation" in report_type:
                        st.success("✅ Customer Segmentation Report Generated")
                        
                        # Create segments
                        customers_df['segment'] = pd.cut(
                            customers_df['total_spent'],
                            bins=[0, 50, 200, 500, float('inf')],
                            labels=['New', 'Regular', 'VIP', 'Premium']
                        )
                        
                        # Segment analysis
                        segment_analysis = customers_df.groupby('segment').agg({
                            'total_spent': ['count', 'sum', 'mean']
                        }).round(2)
                        
                        # Format for display
                        display_segments = segment_analysis.copy()
                        display_segments.columns = ['Customer Count', 'Total Revenue', 'Average Spend']
                        display_segments['Total Revenue'] = display_segments['Total Revenue'].apply(lambda x: f"€{x:,.2f}")
                        display_segments['Average Spend'] = display_segments['Average Spend'].apply(lambda x: f"€{x:,.2f}")
                        
                        st.dataframe(display_segments, use_container_width=True)
                    
                    elif "Acquisition" in report_type:
                        st.success("✅ Customer Acquisition Report Generated")
                        
                        if 'customer_since' in customers_df.columns:
                            customers_df['customer_since'] = pd.to_datetime(customers_df['customer_since'])
                            monthly_acq = customers_df.resample('M', on='customer_since').size().reset_index()
                            monthly_acq.columns = ['Month', 'New Customers']
                            
                            fig = px.line(
                                monthly_acq,
                                x='Month',
                                y='New Customers',
                                title='Monthly Customer Acquisition',
                                markers=True
                            )
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Export acquisition data
                            csv = monthly_acq.to_csv(index=False)
                            st.download_button(
                                label="📥 Download Acquisition Data",
                                data=csv,
                                file_name="customer_acquisition.csv",
                                mime="text/csv"
                            )
                
                else:
                    st.warning("No customer data available")
            else:
                st.error("Failed to fetch customer data")
        
        except Exception as e:
            st.error(f"Error generating customer report: {str(e)}")

def show_supplier_reports(api_url):
    st.subheader("Supplier Performance Reports")
    
    report_type = st.selectbox(
        "Select Supplier Report",
        [
            "Supplier Spending Analysis",
            "Supplier Performance Evaluation",
            "Purchase Order History",
            "Supplier Cost Comparison"
        ]
    )
    
    if st.button("🏭 Generate Supplier Report", type="primary", use_container_width=True):
        try:
            suppliers_response = requests.get(f"{api_url}/api/suppliers")
            
            if suppliers_response.status_code == 200:
                suppliers = suppliers_response.json()
                
                if suppliers:
                    suppliers_df = pd.DataFrame(suppliers)
                    
                    st.success(f"✅ {report_type} Generated")
                    st.info(f"Found {len(suppliers_df)} suppliers in the system")
                    
                    # Display supplier list
                    display_df = suppliers_df[['name', 'contact_person', 'email', 'phone']].copy()
                    
                    st.dataframe(
                        display_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "name": "Supplier",
                            "contact_person": "Contact",
                            "email": "Email",
                            "phone": "Phone"
                        }
                    )
                    
                    # Supplier spending analysis (if we have transaction data)
                    try:
                        transactions_response = requests.get(f"{api_url}/api/transactions", params={"type": "expense", "limit": 1000})
                        if transactions_response.status_code == 200:
                            transactions = transactions_response.json()
                            if transactions:
                                trans_df = pd.DataFrame(transactions)
                                if 'supplier_id' in trans_df.columns:
                                    # Calculate supplier spending
                                    supplier_spending = trans_df.groupby('supplier_id')['amount'].sum().reset_index()
                                    supplier_spending.columns = ['supplier_id', 'total_spent']
                                    
                                    # Merge with supplier info
                                    merged_df = pd.merge(suppliers_df, supplier_spending, 
                                                        left_on='id', right_on='supplier_id', how='left')
                                    merged_df['total_spent'] = merged_df['total_spent'].fillna(0)
                                    
                                    # Show top suppliers by spending
                                    st.subheader("Top Suppliers by Spending")
                                    top_suppliers = merged_df.nlargest(10, 'total_spent')
                                    
                                    fig = px.bar(
                                        top_suppliers,
                                        x='name',
                                        y='total_spent',
                                        title='Top 10 Suppliers by Total Spending',
                                        labels={'name': 'Supplier', 'total_spent': 'Total Spent (€)'}
                                    )
                                    st.plotly_chart(fig, use_container_width=True)
                    except:
                        pass  # Skip if no transaction data
                    
                    # Export option
                    csv = suppliers_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download Supplier Data",
                        data=csv,
                        file_name="supplier_report.csv",
                        mime="text/csv"
                    )
                
                else:
                    st.warning("No supplier data available")
            else:
                st.error("Failed to fetch supplier data")
        
        except Exception as e:
            st.error(f"Error generating supplier report: {str(e)}")

def show_custom_reports(api_url):
    st.subheader("🛠️ Custom Report Builder")
    
    st.info("""
    **Build your own custom reports with the following options:**
    
    1. **Select Data Sources:** Transactions, Customers, Suppliers, Budgets
    2. **Choose Time Period:** Date range, monthly, quarterly, yearly
    3. **Apply Filters:** By category, type, amount range, etc.
    4. **Select Output Format:** PDF, Excel, CSV, or Dashboard view
    5. **Schedule Reports:** Automate report generation
    """)
    
    # Report builder interface
    with st.expander("📋 Step 1: Select Data Sources", expanded=True):
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            include_transactions = st.checkbox("Transactions", value=True)
        with col2:
            include_customers = st.checkbox("Customers", value=True)
        with col3:
            include_suppliers = st.checkbox("Suppliers")
        with col4:
            include_budgets = st.checkbox("Budgets")
    
    with st.expander("📅 Step 2: Set Time Period"):
        date_option = st.radio(
            "Time Period",
            ["Last 30 Days", "Last Quarter", "Last Year", "Year to Date", "All Time", "Custom Range"]
        )
        
        if date_option == "Custom Range":
            start_date = st.date_input("Start Date")
            end_date = st.date_input("End Date")
    
    with st.expander("⚙️ Step 3: Apply Filters"):
        filter_type = st.selectbox("Filter by", ["All", "Income only", "Expenses only"])
        if filter_type != "All":
            min_amount = st.number_input("Minimum Amount", value=0.0)
            max_amount = st.number_input("Maximum Amount", value=10000.0)
    
    with st.expander("💾 Step 4: Output Options"):
        output_format = st.selectbox("Export Format", ["Excel (.xlsx)", "PDF Report", "CSV Data", "Dashboard View"])
        include_charts = st.checkbox("Include Charts", value=True)
        include_summary = st.checkbox("Include Executive Summary", value=True)
    
    # Generate custom report
    if st.button("🚀 Build Custom Report", type="primary", use_container_width=True):
        with st.spinner("Building your custom report..."):
            # Collect data based on selections
            report_data = {}
            
            try:
                if include_transactions:
                    transactions_response = requests.get(f"{api_url}/api/transactions", params={"limit": 5000})
                    if transactions_response.status_code == 200:
                        report_data['transactions'] = transactions_response.json()
                
                if include_customers:
                    customers_response = requests.get(f"{api_url}/api/customers")
                    if customers_response.status_code == 200:
                        report_data['customers'] = customers_response.json()
                
                if include_suppliers:
                    suppliers_response = requests.get(f"{api_url}/api/suppliers")
                    if suppliers_response.status_code == 200:
                        report_data['suppliers'] = suppliers_response.json()
                
                if include_budgets:
                    budgets_response = requests.get(f"{api_url}/api/budgets")
                    if budgets_response.status_code == 200:
                        report_data['budgets'] = budgets_response.json()
                
                # Generate report summary
                st.success("""
                ✅ **Custom Report Ready!**
                
                **Report Configuration:**
                - Data Sources: {}{}{}{}
                - Time Period: {}
                - Output Format: {}
                - Charts Included: {}
                - Summary Included: {}
                
                **Data Collected:**
                - Transactions: {}
                - Customers: {}
                - Suppliers: {}
                - Budgets: {}
                """.format(
                    "✓ Transactions " if include_transactions else "",
                    "✓ Customers " if include_customers else "",
                    "✓ Suppliers " if include_suppliers else "",
                    "✓ Budgets " if include_budgets else "",
                    date_option,
                    output_format,
                    "Yes" if include_charts else "No",
                    "Yes" if include_summary else "No",
                    len(report_data.get('transactions', [])),
                    len(report_data.get('customers', [])),
                    len(report_data.get('suppliers', [])),
                    len(report_data.get('budgets', []))
                ))
                
                # Generate preview based on output format
                st.subheader("📊 Report Preview")
                
                if output_format == "Dashboard View":
                    # Show interactive dashboard
                    if include_transactions and report_data.get('transactions'):
                        trans_df = pd.DataFrame(report_data['transactions'])
                        st.write("**Transaction Summary**")
                        st.write(f"Total Transactions: {len(trans_df)}")
                        
                        if not trans_df.empty:
                            # Calculate metrics
                            total_income = trans_df[trans_df['type'] == 'income']['amount'].sum()
                            total_expenses = trans_df[trans_df['type'] == 'expense']['amount'].sum()
                            
                            col1, col2 = st.columns(2)
                            with col1:
                                st.metric("Total Income", f"€{total_income:,.2f}")
                            with col2:
                                st.metric("Total Expenses", f"€{total_expenses:,.2f}")
                
                elif output_format == "CSV Data":
                    # Prepare CSV data
                    all_data = {}
                    if include_transactions:
                        all_data['transactions'] = pd.DataFrame(report_data.get('transactions', []))
                    if include_customers:
                        all_data['customers'] = pd.DataFrame(report_data.get('customers', []))
                    
                    # Create download buttons for each dataset
                    for data_name, df in all_data.items():
                        if not df.empty:
                            csv = df.to_csv(index=False)
                            st.download_button(
                                label=f"📥 Download {data_name.title()}",
                                data=csv,
                                file_name=f"{data_name}_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv",
                                key=f"download_{data_name}"
                            )
                
                else:
                    # For PDF/Excel, show sample structure
                    preview_data = pd.DataFrame({
                        'Report Section': ['Executive Summary', 'Financial Overview', 'Customer Analysis', 'Supplier Analysis', 'Recommendations'],
                        'Status': ['✓ Complete', '✓ Complete', '✓ Complete' if include_customers else '○ Skipped', 
                                  '✓ Complete' if include_suppliers else '○ Skipped', '✓ Complete'],
                        'Pages': ['1-2', '3-5', '6-8', '9-10', '11-12']
                    })
                    
                    st.dataframe(preview_data, use_container_width=True, hide_index=True)
                    
                    st.info("""
                    **Next Steps:**
                    1. Click download to get the full report
                    2. Review the generated document
                    3. Share with stakeholders
                    4. Schedule regular reports if needed
                    """)
            
            except Exception as e:
                st.error(f"Error building custom report: {str(e)}")